import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name=[]
# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row=line.rstrip().split(",") 
        name.append(row)

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)

find= driver.find_element(By.ID, "input")
find.send_keys("Hunter Hahn")

find = driver.find_element(By.ID,"output")
temp=find.get_attribute("value")

wb = load_workbook("salary.xlsx")
sheet = wb.active

total_salary = 0
for rinda in range(2, sheet.max_row + 1):
    code = sheet['A' + str(rinda)].value
    salary = sheet['B' + str(rinda)].value
    if code == temp:  
        total_salary += int(salary)  

print(total_salary)

wb.close()

driver.quit()
